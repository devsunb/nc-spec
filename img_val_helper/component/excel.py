from openpyxl import Workbook
from openpyxl.drawing.image import Image


class Excel:
    def __init__(self, save_path):
        self.output_wb = Workbook()
        self.save_path = save_path

    def close(self):
        self.output_wb.save(self.save_path)
        self.output_wb.close()

    def set_file_info(self, file_info):
        ws = self.output_wb.create_sheet('개요')
        ws['A1'] = '파일 이름'
        ws['B1'] = file_info.file_name
        ws['A2'] = '촬영 시간'
        ws['B2'] = file_info.scene_time
        ws['A3'] = '슬롯 정보'
        ws['B3'] = file_info.slot
        ws['A4'] = '알고리즘 프로젝트 이름'
        ws['B4'] = file_info.algo_name
        ws['A5'] = '산출물 이름'
        ws['B5'] = file_info.output_name
        ws['A6'] = '/geophysical_data 하위 Variable 개수'
        ws['B6'] = len(file_info.var_list)
        ws['A7'] = '/geophysical_data 하위 Variable 목록'
        ws['B7'] = str(file_info.var_list)

    def set_var_info(self, var_info):
        ws = self.output_wb.create_sheet(var_info.var_name + ' 개요')
        ws['A1'] = 'Variable 이름'
        ws['B1'] = var_info.var_name
        ws['A2'] = '자료형'
        ws['B2'] = str(var_info.dtype)
        ws['A3'] = 'Shape'
        ws['B3'] = str(var_info.shape)
        ws['A4'] = 'Fill value'
        ws['B4'] = var_info.fill_value
        ws['A5'] = '밴드 정보'
        ws['B5'] = var_info.band if var_info.band != 0 else '없음'
        ws['C1'] = '데이터 통계'
        ws['C2'] = 'Average'
        ws['D2'] = var_info.avg
        ws['C3'] = 'Min'
        ws['D3'] = var_info.min
        ws['C4'] = 'Max'
        ws['D4'] = var_info.max
        ws['C5'] = 'NaN ratio'
        ws['D5'] = var_info.nan_ratio
        ws['F1'] = '미리보기'
        ws.add_image(Image(var_info.img_path), 'F2')

    def set_var_data(self, var_name, var_data):
        ws = self.output_wb.create_sheet(var_name)
        for row in var_data:
            ws.append(row.tolist())
